from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from datetime import datetime
import os
from dotenv import load_dotenv
from functools import wraps
from pymongo import MongoClient
import openpyxl
import xlrd
import re

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key')

# MongoDB Connection
MONGO_URI = ""
client = MongoClient(MONGO_URI)
db = client.mess_attendance

# Fixed credentials
ADMIN_USERNAME = ""
ADMIN_PASSWORD = ""

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def clean_enrollment(value):
    """Clean and validate enrollment number."""
    if not value:
        return None
    
    value = str(value).strip()
    
    value = re.sub(r'[^\d.]', '', value)
    
    value = value.split('.')[0]
    
    if len(value) == 8 and value.isdigit():
        return value
    
    return None

def process_excel_file(file, file_ext):
    attendance_data = []
    
    if file_ext == 'xls':
        workbook = xlrd.open_workbook(file_contents=file.read())
        sheet = workbook.sheet_by_index(0)
        
        for row_idx in range(sheet.nrows):
            if row_idx < 4:
                continue
            
            try:
                
                enrollment_cell = sheet.cell(row_idx, 3)
                enrollment = clean_enrollment(enrollment_cell.value)
                
                status_cell = sheet.cell(row_idx, 9)
                status_value = str(status_cell.value).strip().lower()
                
                if not enrollment:
                    continue
                
                status = 'absent' if 'not present' in status_value else 'present'
                
                attendance_data.append({
                    'enrollment': enrollment,
                    'status': status
                })
                
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue
    
    else:  # xlsx files
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=5), start=1):  
            try:
                enrollment_cell = row[3]  
                enrollment = clean_enrollment(enrollment_cell.value)
                
                status_cell = row[9]  
                status_value = str(status_cell.value).strip().lower() if status_cell.value else ''
                
                if not enrollment:
                    continue
                
                status = 'absent' if 'not present' in status_value else 'present'
                
                attendance_data.append({
                    'enrollment': enrollment,
                    'status': status
                })
                
            except Exception as e:
                print(f"Error processing row {row_idx}: {str(e)}")
                continue
    
    return attendance_data



@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('admin'))
        return render_template('login.html', error='Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/admin')
@login_required
def admin():
    return render_template('admin.html')

@app.route('/upload', methods=['POST'])
@login_required
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        if file_ext not in ['xls', 'xlsx']:
            return jsonify({'error': 'Invalid file format. Only .xls and .xlsx files are allowed'}), 400
        
        date = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        attendance_data = process_excel_file(file, file_ext)
        
        if not attendance_data:
            return jsonify({'error': 'No valid data found in file'}), 400
        
        for record in attendance_data:
            record['date'] = date
        
        print(f"Total records to save: {len(attendance_data)}")
        print("First few records:", attendance_data[:5])
        
        db.attendance.insert_many(attendance_data)
        
        return jsonify({'message': f'Upload successful. Processed {len(attendance_data)} records'})
    
    except Exception as e:
        print(f"Error during upload: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/history')
@login_required
def history():
    return render_template('history.html')

@app.route('/api/attendance')
@login_required
def get_attendance():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    try:
        query = {}
        if start_date and end_date:
            query['date'] = {
                '$gte': start_date,
                '$lte': end_date
            }
        
        attendance = list(db.attendance.find(query, {'_id': 0}))
        
        for record in attendance:
            if isinstance(record.get('date'), datetime):
                record['date'] = record['date'].strftime('%Y-%m-%d')
        
        return jsonify(attendance)
    
    except Exception as e:
        print("Error in get_attendance:", str(e))
        return jsonify({'error': str(e)}), 500

@app.route('/absences')
@login_required
def absences():
    return render_template('absences.html')


@app.route('/api/absences')
@login_required
def get_absences():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    consecutive_days = int(request.args.get('consecutive_days', 3))
    
    query = {
        'date': {'$gte': start_date, '$lte': end_date},
        'status': 'absent'
    }
    
    records = list(db.attendance.find(query).sort('date', 1))
    
    absences = {}
    consecutive_absences = []
    
    for record in records:
        enrollment = record['enrollment']
        if enrollment not in absences:
            absences[enrollment] = []
        absences[enrollment].append(record['date'])
    
    for enrollment, dates in absences.items():
        dates.sort()
        consecutive = 1
        start_date = dates[0]
        
        for i in range(1, len(dates)):
            if (datetime.strptime(dates[i], '%Y-%m-%d') - 
                datetime.strptime(dates[i-1], '%Y-%m-%d')).days == 1:
                consecutive += 1
            else:
                if consecutive >= consecutive_days:
                    consecutive_absences.append({
                        'enrollment': enrollment,
                        'start_date': start_date,
                        'end_date': dates[i-1],
                        'days': consecutive
                    })
                consecutive = 1
                start_date = dates[i]
        
        if consecutive >= consecutive_days:
            consecutive_absences.append({
                'enrollment': enrollment,
                'start_date': start_date,
                'end_date': dates[-1],
                'days': consecutive
            })
    
    return jsonify(consecutive_absences)

if __name__ == '__main__':
    app.run(debug=True)